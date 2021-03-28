import json
import jsmin
import re
from decimal import Decimal

from django.core.management.base import BaseCommand, CommandError
from meds.models import ListaHZZO, AtcCode
from openpyxl import load_workbook, Workbook
from openpyxl.cell import column_index_from_string
from optparse import make_option

cmd_write = None

class Command(BaseCommand):
    args = 'excel_file ...>'
    help = 'Exports arc codes to excel file.'
    option_list = BaseCommand.option_list + (
            make_option('-c', '--cfg', dest='cfg_file',
                        help='parser configuration file',
                        default=None),
            )

    def handle(self, *args, **options):
        global cmd_write
        # self.stdout.write('Options %s, args %s' % (options, args))
        if len(args) < 1:
            self.stdout.write('Missing excel_file argument.\n')
            return
        cmd_write = self.stdout.write
        export(args[0], options['cfg_file'])
        cmd_write = None

default_cfg = {
        'sheet_name' : 'wirkstoffe',
        'start_row' : 2,
        'columns' : {
            'opis' : 'C',
            'sifra' : 'D'
            }
        }

def export(ex_file, cfg_file=None):
    cfg = {}
    cfg.update(default_cfg)
    if None != cfg_file:
        f = open(cfg_file)
        b = f.read()
        cfg.update(json.loads(jsmin.jsmin(b)))

    get_opis = create_get_opis(cfg['columns'])
    set_sifra = create_set_sifra(cfg['columns'])

    wb = load_workbook(ex_file)
    sh = wb[cfg['sheet_name']]
    start_row = cfg['start_row']

    for n,r in enumerate(sh.rows[start_row - 1:],start_row):
        opis = get_opis(r)
        rs = AtcCode.objects.filter(opis__iexact=opis)
        if rs.count() > 0:
            set_sifra(sh, n, rs[0].sifra)

    wb.save(ex_file)
    pass

def create_get_opis(cfg):
    column = cfg['opis']
    column_ix = column_index_from_string(column) - 1
    def f(row):
        return row[column_ix].value
    return f

def create_set_sifra(cfg):
    column = cfg['sifra']
    column_ix = column_index_from_string(column)
    def f(sh, row_ix, v):
        c = sh.cell(row=row_ix, column=column_ix)
        c.value = v
    return f

