import datetime
import os.path

from django.http import HttpResponse
from django.shortcuts import HttpResponseRedirect
from django.core.urlresolvers import reverse
from django.db.models import Sum, Q
from django.views.generic import ListView
from django.template.defaultfilters import floatformat

from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.styles import Style, Alignment, Font
from openpyxl.writer.excel import save_virtual_workbook

import settings

from nabava.models import *

# Create your views here.
def naruci(request):
    # create new nabave
    NaruciOdDobavljaca()

    # return nabava_list
    return HttpResponseRedirect(reverse('nabava.all.view'))

class NabavaItemView(ListView):
    nabava_id = None

    def get_queryset(self):
        return NabavaItem.objects.filter(nabava_id=int(self.kwargs['nabava_id']))\
                .values('artikal__ime', 'nabava_id', 'artikal__trziste__naziv',\
                    'artikal__std_kolicina', 'artikal__jedinice',\
                    'artikal__kratica_dobavljaca')\
                .annotate(Sum('artikal__kolicina'))\
                .order_by('artikal__trziste__naziv', 'artikal__ime') 

    def get_context_data(self, **kwargs):
        context = super(NabavaItemView, self).get_context_data(**kwargs)
        context['nabava'] = Nabava.objects.get(id = int(self.kwargs['nabava_id']))

        return context

    def post(self, request, *args, **kwargs):
        n = self.get_context_data(object_list=self.get_queryset(), **kwargs)['nabava']
        n.zabiljezba = request.POST.get('zabiljezba', '')
        n.save()
        return self.get(request, *args, **kwargs)

class NabavaNarudzbaView(ListView):

    def get_queryset(self):
        sk = self.kwargs['std_kolicina']
        if u'None' == sk:
            sk = None
        j = self.kwargs['jedinice']
        if u'None' == j:
            j = None
        return NabavaItem.objects.filter(\
                nabava_id=int(self.kwargs['nabava_id']),\
                artikal__ime=self.kwargs['ime'],\
                artikal__trziste__naziv=self.kwargs['trziste'],\
                artikal__std_kolicina=sk,\
                artikal__jedinice=j)\
                .all()

    def get_context_data(self, **kwargs):
        context = super(NabavaNarudzbaView, self).get_context_data(**kwargs)
        context['nabava'] = Nabava.objects.get(id = int(self.kwargs['nabava_id']))
        context['ime'] = self.kwargs['ime']
        context['trziste'] = self.kwargs['trziste']
        context['std_kolicina'] = self.kwargs['std_kolicina']
        context['jedinice'] = self.kwargs['jedinice']

        return context

def export_nabava(request, nabava_id):

    nabava = Nabava.objects.get(id = int(nabava_id))
    data = make_export_file(nabava)
    if data is not None:
        response = HttpResponse(data, mimetype="application/xlsx")
        response["Content-Disposition"] = "attachment; filename=%s_Konto%d_Nr%s.xlsx" % (nabava.created.date(), nabava.konto.sifra, nabava_id)
        return response
    else:
        return HttpResponse("")

def make_export_file(nabava):

    # collect the data for export
    items =  NabavaItem.objects.filter(nabava_id=nabava.id)\
                .values('artikal__ime', 'nabava_id', 'artikal__trziste__naziv',\
                    'artikal__std_kolicina', 'artikal__jedinice',\
                    'artikal__kratica_dobavljaca')\
                .annotate(Sum('artikal__kolicina'))\
                .order_by('artikal__trziste__naziv', 'artikal__ime') 


    wb = load_workbook(os.path.join(settings.local.ProjectPath, 'Nabava_template.xlsx'))
    sh = wb.active
    #sh.title = 'Bestellung'

    alignment = Alignment(horizontal='center')
    font = Font(bold=True)
    style = Style(alignment=alignment, font=font)

    #sh.append({ 'E' : u'Zu H\u00e4nden %s' % (nabava.konto.kontakt,) })
    c = sh.cell('E1')
    c.value = c.value + str(nabava.konto.sifra)
    #sh.append({})
    #sh.append({ 'E' : u'Bestellung Kd.-Nr.%s' % (nabava.konto.sifra,),\
    #        'G' : nabava.created.date()})
    c = sh.cell('E3')
    c.value = c.value + str(nabava.id)
    sh.cell('G3').value = nabava.created.date()
    #sh.append({})
    #sh.append({\
    #        'A' : 'Nr.',\
    #        'B' : 'Menge',\
    #        'E' : 'Artikelbezeichnung',\
    #        'F' : 'Land',\
    #        'G' : 'Firma'\
    #        })
    for i, it in enumerate(items, 1):
        if it['artikal__trziste__naziv'].startswith('Njema'):
            trz = 'D'
        else:
            trz = it['artikal__trziste__naziv'][:1] 
        sh.append({\
                'A' : i,\
                'B' : it['artikal__kolicina__sum'],\
                'C' : 'x',\
                'D' : (floatformat(it.get('artikal__std_kolicina', '')) or '') + (it.get('artikal__jedinice', '') or ''),\
                'E' : it['artikal__ime'][:40],\
                'F' : trz,\
                'G' : it['artikal__kratica_dobavljaca']\
                })
        c = sh.cell(row=sh.max_row, column=6)
        c.style = style

    return save_virtual_workbook(wb)
    pass

def make_export_file_old(nabava):

    # collect the data for export
    items =  NabavaItem.objects.filter(nabava_id=nabava.id)\
                .values('artikal__ime', 'nabava_id', 'artikal__trziste__naziv',\
                    'artikal__std_kolicina', 'artikal__jedinice',\
                    'artikal__kratica_dobavljaca')\
                .annotate(Sum('artikal__kolicina'))\
                .order_by('artikal__trziste__naziv', 'artikal__ime') 


    wb = Workbook()
    sh = wb.active
    sh.title = 'Bestellung'

    sh.append({ 'E' : u'Zu H\u00e4nden %s' % (nabava.konto.kontakt,) })
    sh.append({})
    sh.append({ 'E' : u'Bestellung Kd.-Nr.%s' % (nabava.konto.sifra,),\
            'G' : nabava.created.date()})
    sh.append({})
    sh.append({\
            'A' : 'Nr.',\
            'B' : 'Menge',\
            'E' : 'Artikelbezeichnung',\
            'F' : 'Land',\
            'G' : 'Firma'\
            })
    for i, it in enumerate(items, 1):
        sh.append({\
                'A' : i,\
                'B' : it['artikal__kolicina__sum'],\
                'C' : 'x',\
                'D' : (floatformat(it.get('artikal__std_kolicina', '')) or '') + (it.get('artikal__jedinice', '') or ''),\
                'E' : it['artikal__ime'],\
                'F' : it['artikal__trziste__naziv'],\
                'G' : it['artikal__kratica_dobavljaca']\
                })

    return save_virtual_workbook(wb)
    pass
